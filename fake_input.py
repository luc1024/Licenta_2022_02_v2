import json
from openpyxl import load_workbook
from faker import Faker
from faker.providers import internet

from env_class import Env


env = Env()
fake = Faker()
fake.add_provider(internet)

wb = load_workbook('fake_samples/fake_sample_input.xlsx')
sh = wb.worksheets[0]

# Fill Email column with fake data
for i in range(env.offset_row + 1, env.offset_row + env.courses_count + 1):
    sh.cell(row=i, column=2).value = fake.email()

with open('fake_samples/fake_items.json', 'r') as f:
    items = json.load(f)
    f.close()

# Fill questions and answers with fake data from fake_items.json
for item in items:
    row = item['course'] + env.offset_row  # course is 1 based
    col = item['original_index'] * env.nr_paragraphs_per_item + env.offset_col + 1  # original_index is 0 based
    sh.cell(row, col).value = item['question']  # question
    for j in range(0, env.answers_per_item):
        sh.cell(row, col + j + 1).value = item['answers'][j]['text']

wb.save('fake_samples/fake_sample_input.xlsx')
