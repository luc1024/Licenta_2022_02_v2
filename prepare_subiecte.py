import os
import openpyxl
import numpy as np


def make_subiecte(dict_zile: dict, nr_comisii: int, nr_subiecte: int, nr_materii: int, items: list, output_folder: str)\
        -> list:
    os.mkdir(output_folder)
    list_zile = list(dict_zile.keys())
    range_comisii = range(ord('A'), ord('A') + nr_comisii)
    materii_zile_comisii = {}
    for i in range(1, nr_materii + 1):
        materii_zile_comisii[i] = []
    subiecte = []
    comisii_itemi = {}
    for comisie in [chr(x) for x in range_comisii]:
        comisii_itemi[comisie] = []
        for zi in list_zile:
            for i in range(nr_subiecte):
                subiect = make_subiect(zi, comisie, i, items, materii_zile_comisii, comisii_itemi)
                subiecte.append(subiect)
    write_frecvente(items, output_folder, len(list_zile), len(range_comisii), nr_subiecte)
    randomize_subiecte(subiecte)
    return subiecte


def write_frecvente(items: list, output_folder: str, nr_zile: int, nr_comisii: int, nr_subiecte: int):
    wb = openpyxl.load_workbook('templates/frecvente_itemi_template.xlsx')
    sheet = wb.worksheets[0]
    for item in items:
        sheet.cell(item['materie'] + 1, item['index_original'] + 1 + 1).value = len(item['zile_comisii'])
    sheet.cell(26, ord('W') - ord('A') + 1).value = str(nr_zile) + ' zile * ' + str(nr_comisii) + ' comisii * ' \
                                                    + str(nr_subiecte) + ' subiecte * 7 itemi = ' \
                                                    + str(nr_zile * nr_comisii * nr_subiecte * 7)
    wb.save(output_folder + 'frecvente_itemi.xlsx')


def randomize_subiecte(subiecte: list) -> list:
    for subiect in subiecte:
        for item in subiect['itemi']:
            item['raspunsuri'] = np.random.permutation(item['raspunsuri']).tolist()
        subiect['itemi'] = np.random.permutation(subiect['itemi']).tolist()
    return subiecte


def make_subiect(zi: str, comisie: str, index: int, items: list, materii_zile_comisii: dict, comisii_itemi: dict)\
        -> dict:
    subiect = {
        'zi': zi,
        'comisie': comisie,
        'index': index,
        'itemi': []
    }

    list_materii = choose_materii(zi, comisie, materii_zile_comisii)
    for materie in list_materii:
        materie_items = list((it for it in items if it['materie'] == materie))
        item = get_best_choices_item(materie_items, zi, comisie, materii_zile_comisii, comisii_itemi)
        subiect['itemi'].append(item)

    return subiect


def choose_materii(zi: str, comisie: str, materii_zile_comisii: dict) -> list:
    ranges_categorii = {1: (0, 15), 2: (15, 18), 3: (18, 21), 4: (21, 23)}
    list_materii = []
    for categorie in ranges_categorii.keys():
        m1 = ranges_categorii[categorie][0]
        m2 = ranges_categorii[categorie][1]
        # alegem materiile care au 1) cei mai putini itemi in (zi, comisie) si
        # 2) cei mai putini itemi selectati per total
        arr = np.array(list(materii_zile_comisii.items())[m1:m2], dtype=object)
        sorted_mzc = sorted(np.array(np.random.permutation(arr)), key=lambda x: (x[1].count((zi, comisie)), len(x[1])))
        if categorie == 1:
            list_materii += list(x[0] for x in sorted_mzc[:4])  # din categoria 1 alegem 4 materii, conform cerintelor
        else:
            list_materii += [sorted_mzc[0].tolist()[0]]  # din celelalte categorii cate o materie, conform cerintelor
    return list_materii


def get_best_choices_item(materie_items: list, zi: str, comisie: str, materii_zile_comisii: dict, comisii_itemi: dict):
    # Telul e sa alegem un item care a fost ales pana acum in cele mai putine (zile, comisii)
    # si sa minimizam nr de itemi care se repeta in aceeasi (zi, comisie)

    # materii_zile comisii = dict{ materie: list[ tuple(zi, comisie) ] }
    # ex: {1: [(L, A), (M, A)], 2:[], 3: [(L, B)], ...}
    # items['zile_comisii'] = list( dict {'zi': zi, 'comisie': comisie} )
    # ex: ({'zi': L, 'comisie': A}, {'zi': M, 'comisie': B})
    # items['comisii'] = list( comisie )
    # ex: (A, C, D)

    # minimul itemilor alesi din materie in comisie (pe toate zilele)
    lowest_materie_comisie: int = min(it['comisii'].count(comisie) for it in materie_items)
    # lista itemilor cel mai putin selectati in comisie, fara cei care au fost selectati in aceeasi zi
    lowest_materie_items = list(it
                                for it in materie_items
                                if it['comisii'].count(comisie) == lowest_materie_comisie
                                and {'zi': zi, 'comisie': comisie} not in it['zile_comisii'])
    # din candidatii optimi pt (zi, comisie) alegem pe cei care au fost in cat mai putine comisii per total
    lowest: int = min(len(it['comisii']) for it in lowest_materie_items)
    best_choices_items = list(it for it in lowest_materie_items if len(it['comisii']) == lowest)
    item = np.random.choice(best_choices_items)

    item['zile_comisii'].append({'zi': zi, 'comisie': comisie})
    item['comisii'].append(comisie)
    materii_zile_comisii[item['materie']].append((zi, comisie))
    comisii_itemi[comisie].append(item)

    return item
