import os
import shutil

import openpyxl
import numpy as np


def make_assignments(dict_exam_days: dict, exam_boards_count: int, assignments_per_board: int, courses_count: int, items: list, output_folder: str)\
        -> list:
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
    os.mkdir(output_folder)
    list_exam_days = list(dict_exam_days.keys())
    range_exam_boards = range(ord('A'), ord('A') + exam_boards_count)
    courses_days_boards = {}
    for i in range(1, courses_count + 1):
        courses_days_boards[i] = []
    assignments = []
    boards_items = {}
    for exam_board in [chr(x) for x in range_exam_boards]:
        boards_items[exam_board] = []
        for day in list_exam_days:
            for i in range(assignments_per_board):
                assignment = make_assignment(day, exam_board, i, items, courses_days_boards, boards_items)
                assignments.append(assignment)
    write_frequencies(items, output_folder, len(list_exam_days), len(range_exam_boards), assignments_per_board)
    randomize_assignments(assignments)
    return assignments


def write_frequencies(items: list, output_folder: str, nr_days: int, exam_boards_count: int, nr_assignments: int):
    wb = openpyxl.load_workbook('templates/frecvente_itemi_template.xlsx')
    sheet = wb.worksheets[0]
    for item in items:
        sheet.cell(item['course'] + 1, item['original_index'] + 1 + 1).value = len(item['days_exam_boards'])
    sheet.cell(26, ord('W') - ord('A') + 1).value = str(nr_days) + ' days * ' + str(exam_boards_count) + ' boards * ' \
                                                    + str(nr_assignments) + ' assignments * 7 items = ' \
                                                    + str(nr_days * exam_boards_count * nr_assignments * 7)
    wb.save(output_folder + 'frecvente_itemi.xlsx')


def randomize_assignments(assignments: list) -> list:
    for assignment in assignments:
        for item in assignment['items']:
            item['answers'] = np.random.permutation(item['answers']).tolist()
        assignment['items'] = np.random.permutation(assignment['items']).tolist()
    return assignments


def make_assignment(day: str, exam_board: str, index: int, items: list, courses_days_boards: dict, boards_items: dict)\
        -> dict:
    assignment = {
        'day': day,
        'exam_board': exam_board,
        'index': index,
        'items': []
    }

    list_courses = choose_courses(day, exam_board, courses_days_boards)
    for course in list_courses:
        course_items = list((it for it in items if it['course'] == course))
        item = get_best_choices_item(course_items, day, exam_board, courses_days_boards, boards_items)
        assignment['items'].append(item)

    return assignment


def choose_courses(day: str, exam_board: str, courses_days_boards: dict) -> list:
    ranges_categories = {1: (0, 15), 2: (15, 18), 3: (18, 21), 4: (21, 23)}
    list_courses = []
    for category in ranges_categories.keys():
        m1 = ranges_categories[category][0]
        m2 = ranges_categories[category][1]
        # alegem materiile care au 1) cei mai putini itemi in (day, exam_board) si
        # 2) cei mai putini itemi selectati per total
        arr = np.array(list(courses_days_boards.items())[m1:m2], dtype=object)
        sorted_mzc = sorted(np.array(np.random.permutation(arr)), key=lambda x: (x[1].count((day, exam_board)), len(x[1])))
        if category == 1:
            list_courses += list(x[0] for x in sorted_mzc[:4])  # din categoria 1 alegem 4 courses, conform cerintelor
        else:
            list_courses += [sorted_mzc[0].tolist()[0]]  # din celelalte categorii cate o materie, conform cerintelor
    return list_courses


def get_best_choices_item(course_items: list, day: str, exam_board: str, courses_days_boards: dict, boards_items: dict):
    # Telul e sa alegem un item care a fost ales pana acum in cele mai putine (zile, comisii)
    # si sa minimizam nr de itemi care se repeta in aceeasi (day, exam_board)

    # courses_days_boards = dict{ course: list[ tuple(day, exam_board) ] }
    # ex: {1: [(L, A), (M, A)], 2:[], 3: [(L, B)], ...}
    # items['days_exam_boards'] = list( dict {'day': day, 'exam_board': exam_board} )
    # ex: ({'day': L, 'exam_board': A}, {'day': M, 'exam_board': B})
    # items['exam_boards'] = list( exam_board )
    # ex: (A, C, D)

    # minimul itemilor alesi din course in exam_board (pe toate zilele)
    lowest_course_exam_board: int = min(it['exam_boards'].count(exam_board) for it in course_items)
    # lista itemilor cel mai putin selectati in exam_board, fara cei care au fost selectati in aceeasi day
    lowest_course_items = list(it
                                for it in course_items
                                if it['exam_boards'].count(exam_board) == lowest_course_exam_board
                                and {'day': day, 'exam_board': exam_board} not in it['days_exam_boards'])
    # din candidatii optimi pt (day, exam_board) alegem pe cei care au fost in cat mai putine comisii per total
    lowest: int = min(len(it['exam_boards']) for it in lowest_course_items)
    best_choices_items = list(it for it in lowest_course_items if len(it['exam_boards']) == lowest)
    item = np.random.choice(best_choices_items)

    item['days_exam_boards'].append({'day': day, 'exam_board': exam_board})
    item['exam_boards'].append(exam_board)
    courses_days_boards[item['course']].append((day, exam_board))
    boards_items[exam_board].append(item)

    return item
